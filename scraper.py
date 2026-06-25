import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL = "https://books.toscrape.com/"

RATINGS = {
    "One": "★☆☆☆☆",
    "Two": "★★☆☆☆",
    "Three": "★★★☆☆",
    "Four": "★★★★☆",
    "Five": "★★★★★"
}

RATINGS_VALUE = {
    "One": 1,
    "Two": 2,
    "Three": 3,
    "Four": 4,
    "Five": 5
}

def get_category(book_url):
    response = requests.get(book_url)
    response.encoding = "utf-8"
    soup = BeautifulSoup(response.text, "html.parser")
    breadcrumb = soup.find("ul", class_="breadcrumb")
    if breadcrumb:
        return breadcrumb.find_all("li")[2].text.strip()
    return "Unknown"

def get_menu():
    print("=" * 40)
    print("        📚 BOOKS WEB SCRAPER")
    print("=" * 40)

    # Number of pages
    while True:
        try:
            pages = int(input("📄 How many pages do you want to scrape? (1-50): "))
            if 1 <= pages <= 50:
                break
            print("Please enter a number between 1 and 50.")
        except ValueError:
            print("Please enter a valid number.")

    # Minimum rating
    while True:
        try:
            min_rating = int(input("⭐ Minimum rating filter (1-5): "))
            if 1 <= min_rating <= 5:
                break
            print("Please enter a number between 1 and 5.")
        except ValueError:
            print("Please enter a valid number.")

    # Category filter
    category_filter = input("📂 Filter by category? (leave empty for all): ").strip().lower()

    return pages, min_rating, category_filter

# Show menu and get user choices
pages, min_rating, category_filter = get_menu()

books = []

print("\n⏳ Scraping pages...")

for page_number in range(1, pages + 1):

    url = BASE_URL if page_number == 1 else f"{BASE_URL}catalogue/page-{page_number}.html"

    response = requests.get(url)
    response.encoding = "utf-8"
    soup = BeautifulSoup(response.text, "html.parser")

    for book in soup.find_all("article", class_="product_pod"):
        title = book.find("h3").find("a")["title"]
        price = book.find("p", class_="price_color").text
        rating_word = book.find("p", class_="star-rating")["class"][1]
        rating = RATINGS[rating_word]
        rating_value = RATINGS_VALUE[rating_word]

        # Skip books below minimum rating
        if rating_value < min_rating:
            continue

        book_path = book.find("h3").find("a")["href"].replace("../", "")
        if not book_path.startswith("catalogue/"):
            book_path = f"catalogue/{book_path}"
        book_url = f"{BASE_URL}{book_path}"
        category = get_category(book_url)

        # Skip books that don't match category filter
        if category_filter and category_filter not in category.lower():
            continue

        books.append([title, price, rating, category])
        print(f"  📖 {title[:50]} — {rating} — {category}")

    print(f"✅ Page {page_number}/{pages} done\n")

# Create and style the Excel file
wb = Workbook()
ws = wb.active
ws.title = "Books"

header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for col, header in enumerate(["Title", "Price", "Rating", "Category"], start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for book in books:
    ws.append(book)

if books:
    for col in ws.columns:
        max_width = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_width + 4
    wb.save("books.xlsx")
    print(f"\n🏆 Done! {len(books)} books saved to books.xlsx")
else:
    print("\n⚠️ No books found with those filters. Try different options.")